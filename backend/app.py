import sys
import os

# Add pyaadhaar library path
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..',   'pylib'))

import cv2
import numpy as np
from flask import Flask, request, jsonify
from flask_cors import CORS
from pyaadhaar.decode import AadhaarSecureQr, AadhaarOldQr

app = Flask(__name__)
CORS(app)


def read_qr_from_image(image_bytes):
    """Extract raw QR code data from image bytes using OpenCV QR detector."""
    np_arr = np.frombuffer(image_bytes, np.uint8)
    img = cv2.imdecode(np_arr, cv2.IMREAD_COLOR)
    if img is None:
        return None, "Could not read image"

    detector = cv2.QRCodeDetector()

    # Try on original image
    data, _, _ = detector.detectAndDecode(img)
    if data:
        return data, None

    # Try on grayscale
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    data, _, _ = detector.detectAndDecode(gray)
    if data:
        return data, None

    # Try on upscaled image for small QR codes
    h, w = img.shape[:2]
    scale = max(1, 2000 // max(h, w))
    if scale > 1:
        upscaled = cv2.resize(img, (w * scale, h * scale), interpolation=cv2.INTER_CUBIC)
        data, _, _ = detector.detectAndDecode(upscaled)
        if data:
            return data, None

    return None, "No QR code found in image"


def is_secure_qr(data):
    """Returns True if it's a new Secure QR (all digits), False if old QR (XML)."""
    try:
        int(data)
        return True
    except ValueError:
        return False


@app.route('/api/decode', methods=['POST'])
def decode_aadhaar():
    if 'image' not in request.files:
        return jsonify({'error': 'No image file provided'}), 400

    file = request.files['image']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400

    allowed_extensions = {'png', 'jpg', 'jpeg', 'gif', 'bmp', 'webp'}
    ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else ''
    if ext not in allowed_extensions:
        return jsonify({'error': 'Unsupported file type'}), 400

    image_bytes = file.read()
    qr_data, error = read_qr_from_image(image_bytes)

    if error:
        return jsonify({'error': error}), 422

    try:
        if is_secure_qr(qr_data):
            obj = AadhaarSecureQr(int(qr_data))
            decoded = obj.decodeddata()
            qr_type = 'Secure QR (New)'

            # Check for photo
            has_image = False
            image_b64 = None
            try:
                has_image = obj.isImage()
                if has_image:
                    from PIL import Image
                    import base64
                    from io import BytesIO
                    pil_img = obj.image()
                    buf = BytesIO()
                    pil_img.save(buf, format='JPEG')
                    image_b64 = base64.b64encode(buf.getvalue()).decode('utf-8')
            except Exception:
                has_image = False

            return jsonify({
                'success': True,
                'qr_type': qr_type,
                'data': decoded,
                'has_image': has_image,
                'photo': image_b64
            })

        else:
            obj = AadhaarOldQr(qr_data)
            decoded = obj.decodeddata()
            return jsonify({
                'success': True,
                'qr_type': 'Old QR',
                'data': decoded,
                'has_image': False,
                'photo': None
            })

    except Exception as e:
        return jsonify({'error': f'Decoding failed: {str(e)}'}), 500


@app.route('/api/compress-pdf', methods=['POST'])
def compress_pdf():
    if 'pdf' not in request.files:
        return jsonify({'error': 'No PDF file provided'}), 400

    file = request.files['pdf']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400

    ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else ''
    if ext != 'pdf':
        return jsonify({'error': 'Only PDF files are supported'}), 400

    # mode: 'percent' (0-100 quality) or 'size' (target MB)
    mode = request.form.get('mode', 'percent')
    try:
        if mode == 'percent':
            quality = max(5, min(95, int(request.form.get('quality', 60))))
            target_bytes = None
        else:
            target_mb = float(request.form.get('target_mb', 1.0))
            target_bytes = int(target_mb * 1024 * 1024)
            quality = 60  # starting point for binary search
    except (ValueError, TypeError):
        return jsonify({'error': 'Invalid compression parameters'}), 400

    pdf_bytes = file.read()
    original_size = len(pdf_bytes)

    try:
        import fitz
        from io import BytesIO

        def compress_at_quality(q):
            src = fitz.open(stream=BytesIO(pdf_bytes), filetype='pdf')
            dst = fitz.open()
            for page in src:
                # Render page to pixmap and reinsert as compressed image-based page
                mat = fitz.Matrix(1.5, 1.5)
                pix = page.get_pixmap(matrix=mat, alpha=False)
                img_bytes = pix.tobytes('jpeg', jpg_quality=q)
                img_pdf_bytes = fitz.open('pdf', fitz.open(
                    stream=BytesIO(img_bytes), filetype='jpeg'
                ).convert_to_pdf())
                dst.insert_pdf(img_pdf_bytes)
            buf = BytesIO()
            dst.save(buf, garbage=4, deflate=True)
            dst.close()
            src.close()
            buf.seek(0)
            return buf.read()

        if mode == 'percent':
            compressed = compress_at_quality(quality)
        else:
            # Binary search for quality that gets close to target_bytes
            lo, hi = 5, 95
            compressed = compress_at_quality(lo)  # worst case
            for _ in range(8):
                mid = (lo + hi) // 2
                candidate = compress_at_quality(mid)
                if len(candidate) <= target_bytes:
                    compressed = candidate
                    lo = mid + 1
                else:
                    hi = mid - 1
                if lo > hi:
                    break

        compressed_size = len(compressed)
        savings = round((1 - compressed_size / original_size) * 100, 1)

        base_name = file.filename.rsplit('.', 1)[0]
        from flask import send_file
        response = send_file(
            BytesIO(compressed),
            as_attachment=True,
            download_name=f'{base_name}_compressed.pdf',
            mimetype='application/pdf'
        )
        response.headers['X-Original-Size'] = str(original_size)
        response.headers['X-Compressed-Size'] = str(compressed_size)
        response.headers['X-Savings-Percent'] = str(savings)
        response.headers['Access-Control-Expose-Headers'] = 'X-Original-Size,X-Compressed-Size,X-Savings-Percent'
        return response

    except Exception as e:
        return jsonify({'error': f'Compression failed: {str(e)}'}), 500


@app.route('/api/pdf-to-excel', methods=['POST'])
def pdf_to_excel():
    if 'pdf' not in request.files:
        return jsonify({'error': 'No PDF file provided'}), 400

    file = request.files['pdf']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400

    ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else ''
    if ext != 'pdf':
        return jsonify({'error': 'Only PDF files are supported'}), 400

    pdf_bytes = file.read()

    try:
        import pdfplumber
        import openpyxl
        from io import BytesIO

        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # remove default sheet, add per-page sheets

        with pdfplumber.open(BytesIO(pdf_bytes)) as pdf:
            if len(pdf.pages) == 0:
                return jsonify({'error': 'PDF has no pages'}), 422

            for page_num, page in enumerate(pdf.pages, start=1):
                ws = wb.create_sheet(title=f'Page {page_num}')
                tables = page.extract_tables()

                if tables:
                    row_idx = 1
                    for table in tables:
                        for row in table:
                            for col_idx, cell in enumerate(row, start=1):
                                ws.cell(row=row_idx, column=col_idx, value=cell or '')
                            row_idx += 1
                        row_idx += 1  # blank row between tables
                else:
                    # No tables — extract raw text into column A
                    text = page.extract_text() or ''
                    for row_idx, line in enumerate(text.splitlines(), start=1):
                        ws.cell(row=row_idx, column=1, value=line)

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        base_name = file.filename.rsplit('.', 1)[0]
        from flask import send_file
        return send_file(
            buf,
            as_attachment=True,
            download_name=f'{base_name}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        return jsonify({'error': f'Conversion failed: {str(e)}'}), 500


@app.route('/api/pdf-to-word', methods=['POST'])
def pdf_to_word():
    if 'pdf' not in request.files:
        return jsonify({'error': 'No PDF file provided'}), 400

    file = request.files['pdf']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400

    ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else ''
    if ext != 'pdf':
        return jsonify({'error': 'Only PDF files are supported'}), 400

    pdf_bytes = file.read()

    try:
        import tempfile
        import os
        from io import BytesIO
        from flask import send_file
        from pdf2docx import Converter

        with tempfile.TemporaryDirectory() as tmpdir:
            pdf_path = os.path.join(tmpdir, 'input.pdf')
            docx_path = os.path.join(tmpdir, 'output.docx')

            with open(pdf_path, 'wb') as f:
                f.write(pdf_bytes)

            cv = Converter(pdf_path)
            cv.convert(docx_path)
            cv.close()

            with open(docx_path, 'rb') as f:
                docx_bytes = f.read()

        base_name = file.filename.rsplit('.', 1)[0]
        return send_file(
            BytesIO(docx_bytes),
            as_attachment=True,
            download_name=f'{base_name}.docx',
            mimetype='application/vnd.openxmlformats-officedocument.wordprocessingml.document'
        )

    except Exception as e:
        return jsonify({'error': f'Conversion failed: {str(e)}'}), 500


@app.route('/api/pdf-to-jpg', methods=['POST'])
def pdf_to_jpg():
    if 'pdf' not in request.files:
        return jsonify({'error': 'No PDF file provided'}), 400

    file = request.files['pdf']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400

    ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else ''
    if ext != 'pdf':
        return jsonify({'error': 'Only PDF files are supported'}), 400

    pdf_bytes = file.read()

    try:
        import fitz  # pymupdf
        import base64
        from io import BytesIO

        doc = fitz.open(stream=BytesIO(pdf_bytes), filetype='pdf')

        if doc.needs_pass:
            # Try blank password for owner-encrypted PDFs with no user password
            if not doc.authenticate(''):
                doc.close()
                return jsonify({'error': 'PDF is password-protected and cannot be converted'}), 422

        pages = []
        mat = fitz.Matrix(2, 2)  # 2x zoom (~150 dpi)
        for i, page in enumerate(doc):
            pix = page.get_pixmap(matrix=mat)
            img_bytes = pix.tobytes('jpeg')
            b64 = base64.b64encode(img_bytes).decode('utf-8')
            pages.append({'page': i + 1, 'image': b64})
        doc.close()

        return jsonify({'success': True, 'pages': pages, 'total': len(pages)})

    except Exception as e:
        return jsonify({'error': f'Conversion failed: {str(e)}'}), 500


@app.route('/api/health', methods=['GET'])
def health():
    return jsonify({'status': 'ok'})


if __name__ == '__main__':
    app.run(debug=True, port=5000)
